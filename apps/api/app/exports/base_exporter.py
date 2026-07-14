from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.exports.constants import ExportConstants
from app.exports.styles import ExportStyles


class BaseExporter:
    """Base class for all QABook Excel exporters."""

    def __init__(self):
        self.workbook = Workbook()
        self.worksheet: Worksheet = self.workbook.active
        self.current_row = 1

    def create_document(self, sheet_name: str, document_title: str):
        """Initialize workbook and create the common document layout."""

        self.worksheet.title = sheet_name

        self.set_default_layout()
        self.add_document_header(document_title)

    def set_default_layout(self):
        """Apply worksheet defaults."""
    
        column_widths = {
            "A": 22,
            "B": 24,
            "C": 16,
            "D": 16,
            "E": 60,
            "F": 25,
            "G": 25,
            "H": 20,
        }
    
        for column, width in column_widths.items():
            self.worksheet.column_dimensions[column].width = width
    
        # Landscape printing
        self.worksheet.page_setup.orientation = "landscape"
    
        # Fit all columns on one page
        self.worksheet.page_setup.fitToWidth = 1
        self.worksheet.page_setup.fitToHeight = 0
    
        # Center the document when printing
        self.worksheet.print_options.horizontalCentered = True
    
        # Freeze header row
        self.worksheet.freeze_panes = "A12"

    def add_document_header(self, document_title: str):
        """Add the QABook document title."""

        # QABook Title
        self.worksheet.merge_cells("A1:H1")
        cell = self.worksheet["A1"]
        cell.value = ExportConstants.APP_NAME
        cell.font = ExportStyles.TITLE_FONT
        cell.alignment = ExportStyles.CENTER_ALIGNMENT

        # Document Title
        self.worksheet.merge_cells("A2:H2")
        cell = self.worksheet["A2"]
        cell.value = document_title
        cell.font = ExportStyles.DOCUMENT_TITLE_FONT
        cell.alignment = ExportStyles.CENTER_ALIGNMENT

        self.current_row = 4


    def add_project_information(self, metadata: dict):
        """Create the Project Information section."""

        ws = self.worksheet
        row = self.current_row

        # Section Header
        ws.merge_cells(f"A{row}:H{row}")

        cell = ws[f"A{row}"]
        cell.value = "Project Information"
        cell.font = ExportStyles.SECTION_HEADER_FONT
        cell.fill = ExportStyles.METADATA_FILL

        row += 1

        for label, value in metadata.items():

            label_cell = ws[f"A{row}"]
            label_cell.value = label
            label_cell.font = ExportStyles.TABLE_HEADER_FONT
            label_cell.fill = ExportStyles.TABLE_HEADER_FILL
            label_cell.border = ExportStyles.THIN_BORDER

            value_cell = ws[f"B{row}"]
            value_cell.value = value
            value_cell.font = ExportStyles.NORMAL_FONT
            value_cell.border = ExportStyles.THIN_BORDER

            row += 1

        self.current_row = row + 1

    def set_column_widths(
        self,
        widths: dict[str, float],
    ):
        """Set custom column widths for a worksheet."""
    
        for column, width in widths.items():
            self.worksheet.column_dimensions[column].width = width

    def get_workbook(self):
        return self.workbook