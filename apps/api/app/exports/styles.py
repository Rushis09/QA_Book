from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExportStyles:
    """Shared styles for all QABook Excel exports."""

    # -------------------------
    # Fonts
    # -------------------------

    TITLE_FONT = Font(
        name="Calibri",
        size=16,
        bold=True,
    )

    DOCUMENT_TITLE_FONT = Font(
        name="Calibri",
        size=14,
        bold=True,
    )

    SECTION_HEADER_FONT = Font(
        name="Calibri",
        size=12,
        bold=True,
    )

    TABLE_HEADER_FONT = Font(
        name="Calibri",
        size=11,
        bold=True,
        color="FFFFFF",
    )

    NORMAL_FONT = Font(
        name="Calibri",
        size=11,
    )

    # -------------------------
    # Alignments
    # -------------------------

    CENTER_ALIGNMENT = Alignment(
        horizontal="center",
        vertical="center",
    )

    LEFT_ALIGNMENT = Alignment(
        horizontal="left",
        vertical="center",
    )

    WRAP_ALIGNMENT = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    # -------------------------
    # Borders
    # -------------------------

    _THIN_SIDE = Side(
        border_style="thin",
        color="000000",
    )

    THIN_BORDER = Border(
        left=_THIN_SIDE,
        right=_THIN_SIDE,
        top=_THIN_SIDE,
        bottom=_THIN_SIDE,
    )

    # -------------------------
    # Cell Fills
    # -------------------------

    TABLE_HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    METADATA_FILL = PatternFill(
        fill_type="solid",
        fgColor="EDEDED",
    )